import pandas as pd
from datetime import datetime
from dateutil import parser
import json
from pathlib import Path
from openpyxl import load_workbook

# -------------------------------------------------------
# CONFIG
# -------------------------------------------------------

COMPANY_JSON_PATH = Path(__file__).resolve().parent / "company.json"

if not COMPANY_JSON_PATH.exists():
    raise FileNotFoundError(f"company.json not found at: {COMPANY_JSON_PATH}")

with COMPANY_JSON_PATH.open("r", encoding="utf-8") as f:
    json_db = json.load(f)

MANDATORY_HEADERS = [
    "Employee Id",
    "Options Granted",
    "Plan Name",
    "Date Of Grant",
    "Grant Price",
    "Vesting Template",
    "Vesting Date Type"
]

VESTING_DATE_TYPE_OPTS = ["GrantDate", "CustomDate", "EmployeeJoiningDate"]

ACTUAL_VESTING_DAY_OPTS = [
    "SAME_DAY",
    "NEXT_DAY",
    "PREVIOUS_DAY",
    "STARTING_OF_MONTH",
    "END_OF_MONTH"
]

# Date format mapping: column name -> Excel format string
DATE_FORMATS = {
    "Date Of Grant": "D/M/YYYY",
    "Vesting Date": "d-m-yyyy"
}

# -------------------------------------------------------
# DATE PARSER (RETURNS DATETIME OBJECT)
# -------------------------------------------------------

def parse_date(value):

    if pd.isna(value) or str(value).strip() == "":
        raise ValueError("Date is empty")

    # Excel serial
    if isinstance(value, (int, float)):
        return pd.to_datetime(value, unit="D", origin="1899-12-30")

    # Already datetime
    if isinstance(value, (pd.Timestamp, datetime)):
        return value

    # String
    s = str(value).strip().replace("/", "-").replace(".", "-")

    try:
        return pd.to_datetime(s, dayfirst=True)
    except:
        try:
            return parser.parse(s, dayfirst=True)
        except:
            raise ValueError(f"Invalid date format: {value}")

# -------------------------------------------------------
# IN-PLACE DATE VALIDATION AND FORMATTING
# -------------------------------------------------------

def validate_and_format_dates_inplace(file_path, company_key, json_db):
    """
    Validates the Excel file and formats dates in-place without creating a new file.
    This preserves all existing formatting and data types.
    
    Returns:
        tuple: (error_report, success) where error_report is None if successful
    """
    
    error_report = {
        "file_status": "Valid",
        "file_errors": [],
        "row_errors": {}
    }
    
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Get headers from first row
    headers = {}
    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        header = col[0].value
        if header:
            headers[str(header).strip()] = col_idx
    
    # Check mandatory headers
    missing = [c for c in MANDATORY_HEADERS if c not in headers]
    if missing:
        error_report["file_status"] = "Has Errors"
        error_report["file_errors"] = [
            f"Missing mandatory column header: {c}" for c in missing
        ]
        return error_report, False
    
    # Get company config
    try:
        valid_plans = set(json_db[company_key]["plan_names"])
        valid_templates = set(json_db[company_key]["vesting_templates"])
    except:
        return {
            "file_status": "Fatal Error",
            "file_errors": ["Company Key not found in JSON DB"],
            "row_errors": {}
        }, False
    
    # Validate each row
    for row_idx in range(2, ws.max_row + 1):
        row_err = {}
        
        # Employee Id
        emp_col = headers.get("Employee Id")
        if emp_col:
            emp = ws.cell(row=row_idx, column=emp_col).value
            emp = str(emp).strip() if emp else ""
            if not emp or emp.lower() == "nan" or emp.lower() == "none":
                row_err["Employee Id"] = "Field is empty."
        
        # Options Granted
        opt_col = headers.get("Options Granted")
        if opt_col:
            try:
                v = ws.cell(row=row_idx, column=opt_col).value
                if v is None:
                    raise ValueError("Empty")
                v = float(v)
                if v <= 0 or not v.is_integer():
                    row_err["Options Granted"] = "Must be whole number > 0."
            except:
                row_err["Options Granted"] = "Invalid number."
        
        # Plan Name
        plan_col = headers.get("Plan Name")
        if plan_col:
            plan = ws.cell(row=row_idx, column=plan_col).value
            plan = str(plan).strip() if plan else ""
            if not plan:
                row_err["Plan Name"] = "Field is empty."
            elif plan not in valid_plans:
                row_err["Plan Name"] = "Invalid Plan."
        
        # Date Of Grant - validate and format
        grant_date_col = headers.get("Date Of Grant")
        if grant_date_col:
            cell = ws.cell(row=row_idx, column=grant_date_col)
            try:
                if cell.value is None:
                    raise ValueError("Date is empty")
                # If it's already a datetime, just apply format
                if isinstance(cell.value, datetime):
                    cell.number_format = DATE_FORMATS["Date Of Grant"]
                else:
                    # Parse and set the value
                    parsed = parse_date(cell.value)
                    cell.value = parsed
                    cell.number_format = DATE_FORMATS["Date Of Grant"]
            except ValueError as e:
                row_err["Date Of Grant"] = str(e)
        
        # Grant Price
        price_col = headers.get("Grant Price")
        if price_col:
            try:
                v = ws.cell(row=row_idx, column=price_col).value
                if v is None:
                    raise ValueError("Empty")
                float(v)
            except:
                row_err["Grant Price"] = "Must be number."
        
        # Vesting Template
        template_col = headers.get("Vesting Template")
        if template_col:
            template = ws.cell(row=row_idx, column=template_col).value
            template = str(template).strip() if template else ""
            if not template:
                row_err["Vesting Template"] = "Field is empty."
            elif template not in valid_templates:
                row_err["Vesting Template"] = "Invalid Template."
        
        # Vesting Date Type
        vtype_col = headers.get("Vesting Date Type")
        vtype = None
        if vtype_col:
            vtype = ws.cell(row=row_idx, column=vtype_col).value
            vtype = str(vtype).strip() if vtype else ""
            if vtype not in VESTING_DATE_TYPE_OPTS:
                row_err["Vesting Date Type"] = "Invalid Type."
        
        # Vesting Date - validate and format if CustomDate
        if vtype == "CustomDate":
            vdate_col = headers.get("Vesting Date")
            if vdate_col:
                cell = ws.cell(row=row_idx, column=vdate_col)
                try:
                    if cell.value is None:
                        raise ValueError("Date is empty")
                    # If it's already a datetime, just apply format
                    if isinstance(cell.value, datetime):
                        cell.number_format = DATE_FORMATS["Vesting Date"]
                    else:
                        # Parse and set the value
                        parsed = parse_date(cell.value)
                        cell.value = parsed
                        cell.number_format = DATE_FORMATS["Vesting Date"]
                except:
                    row_err["Vesting Date"] = "Invalid/Empty Date."
        
        # Actual Vesting Day
        act_col = headers.get("Actual Vesting Day")
        if act_col:
            act = ws.cell(row=row_idx, column=act_col).value
            act = str(act).strip() if act else ""
            if act and act not in ACTUAL_VESTING_DAY_OPTS:
                row_err["Actual Vesting Day"] = "Invalid Option."
        
        if row_err:
            error_report["row_errors"][str(row_idx)] = row_err
            error_report["file_status"] = "Has Errors"
    
    # If there are errors, don't save
    if error_report["file_errors"] or error_report["row_errors"]:
        return error_report, False
    
    # Save the file with all validations and formatting applied
    wb.save(file_path)
    return None, True


# -------------------------------------------------------
# LEGACY FUNCTION (for backward compatibility)
# -------------------------------------------------------

def process_and_validate_excel(df, company_key, json_db):
    """
    Legacy function - kept for backward compatibility.
    Note: This returns a DataFrame which will be saved separately.
    Use validate_and_format_dates_inplace() for in-place editing.
    """
    
    error_report = {
        "file_status": "Valid",
        "file_errors": [],
        "row_errors": {}
    }

    # Normalize headers
    df.columns = [str(c).strip() for c in df.columns]
    cols = df.columns.tolist()

    # Mandatory headers
    missing = [c for c in MANDATORY_HEADERS if c not in cols]
    if missing:
        error_report["file_status"] = "Has Errors"
        error_report["file_errors"] = [
            f"Missing mandatory column header: {c}" for c in missing
        ]

    # Company config
    try:
        valid_plans = set(json_db[company_key]["plan_names"])
        valid_templates = set(json_db[company_key]["vesting_templates"])
    except:
        return {
            "file_status": "Fatal Error",
            "file_errors": ["Company Key not found in JSON DB"],
            "row_errors": {}
        }, None

    # Row validations
    for idx, row in df.iterrows():
        excel_row = idx + 2
        row_err = {}

        # Employee Id
        emp = str(row.get("Employee Id", "")).strip()
        if not emp or emp.lower() == "nan":
            row_err["Employee Id"] = "Field is empty."

        # Options Granted
        try:
            v = float(row.get("Options Granted"))
            if v <= 0 or not v.is_integer():
                row_err["Options Granted"] = "Must be whole number > 0."
        except:
            row_err["Options Granted"] = "Invalid number."

        # Plan Name
        plan = str(row.get("Plan Name", "")).strip()
        if not plan:
            row_err["Plan Name"] = "Field is empty."
        elif plan not in valid_plans:
            row_err["Plan Name"] = "Invalid Plan."

        # Date Of Grant
        try:
            parsed_date = parse_date(row.get("Date Of Grant"))
            df.at[idx, "Date Of Grant"] = parsed_date
        except ValueError as e:
            row_err["Date Of Grant"] = str(e)

        # Grant Price
        try:
            float(row.get("Grant Price"))
        except:
            row_err["Grant Price"] = "Must be number."

        # Vesting Template
        template = str(row.get("Vesting Template", "")).strip()
        if not template:
            row_err["Vesting Template"] = "Field is empty."
        elif template not in valid_templates:
            row_err["Vesting Template"] = "Invalid Template."

        # Vesting Date Type
        vtype = str(row.get("Vesting Date Type", "")).strip()
        if vtype not in VESTING_DATE_TYPE_OPTS:
            row_err["Vesting Date Type"] = "Invalid Type."

        # Vesting Date
        if vtype == "CustomDate":
            try:
                parsed_vesting_date = parse_date(row.get("Vesting Date"))
                df.at[idx, "Vesting Date"] = parsed_vesting_date
            except:
                row_err["Vesting Date"] = "Invalid/Empty Date."

        # Actual Vesting Day
        act = str(row.get("Actual Vesting Day", "")).strip()
        if act and act not in ACTUAL_VESTING_DAY_OPTS:
            row_err["Actual Vesting Day"] = "Invalid Option."

        if row_err:
            error_report["row_errors"][str(excel_row)] = row_err
            error_report["file_status"] = "Has Errors"

    if error_report["file_errors"] or error_report["row_errors"]:
        return error_report, None

    # Convert date columns to proper datetime dtype
    for col in DATE_FORMATS.keys():
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    return None, df